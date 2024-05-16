import shutil
import tempfile
import uuid
from pathlib import Path

import formulas
from openpyxl import load_workbook


def it1():
    working_file = Path(tempfile.gettempdir(), f'{uuid.uuid4()!s}.xlsm').absolute()
    shutil.copyfile('01d-jedi-geothermal-model-rel-gt12-23-16.xlsm', working_file)
    wb = load_workbook(working_file)
    project_data = wb['ProjectData']
    print(project_data)
    print(project_data['B16'])

    results = wb['Summary Results']
    print(results['B28'].value)

    a1 = wb['Calculations']['AG146'].value
    print(a1)
    for cell in a1[1:].split('+'):
        print(wb['Calculations'][cell].value)


def with_formulas():
    working_file = Path(tempfile.gettempdir(), f'{uuid.uuid4()!s}.xlsm').absolute()
    shutil.copyfile('01d-jedi-geothermal-model-rel-gt12-23-16.xlsm', working_file)

    xl_model = formulas.ExcelModel().loads(str(working_file)).finish()
    xl_model.calculate()
    xl_model.write(dirpath=working_file.parent)

    wb = load_workbook(working_file)
    project_data = wb['ProjectData']
    print(project_data)
    print(project_data['B16'])

    results = wb['Summary Results']
    print(results['B28'].value)

    a1 = wb['Calculations']['AG146'].value
    print(a1)
    for cell in a1[1:].split('+'):
        print(wb['Calculations'][cell].value)


if __name__ == '__main__':
    # it1()
    with_formulas()
